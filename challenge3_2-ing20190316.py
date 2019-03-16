# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def combine():
    wb1 = load_workbook("courses.xlsx")
    ws1 = wb1['students']
    ws2 = wb1['time']
    ws3 = wb1.copy_worksheet(ws1)
    ws3.title = 'combine'

    lr1 = list(ws1.rows)
    lr2 = list(ws2.rows)
    row_num_ws1 = len(lr1)
    row_num_ws2 = len(lr2)

    i = 1
    while i < row_num_ws1:
        j = 1
        course_name = lr1[i][1].value
        while j < row_num_ws2:
            if course_name == lr2[j][1].value :
                ws3.cell(row=i+1,column=4,value=lr2[j][2].value)
                break
            j += 1
        i += 1

    ws3.cell(row=1,column=4,value=lr2[0][2].value)  #biao tou filed
   
    wb1.save("courses.xlsx")

def split():
    wb_all= load_workbook("courses.xlsx")
    ws_combine = wb_all['combine']
    lr_combine = list(ws_combine.rows)
    row_num_ws_combine = len(lr_combine)

    years_set = set()
    i=1
    while i < row_num_ws_combine:
        years_set.add((lr_combine[i][0].value).year)
        i += 1
    years_list = list(years_set)

    i=0
    years_num = len(years_list)
    wb_dict = {}
    while i < years_num:
        key = years_list[i]
        wbnew = Workbook(str(years_list[i],write_only=False) 
        wb_dict[key] = wbnew
        i += 1

    i = 2
    while i < row_num_ws_combine:
        #row_range_combine1 = ws_combine[i]
        j = 1
        if lr_combine[i][0].value == 










    

if __name__ == '__main__':
    combine()




